import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load extracted bid data
with open('bid_data.json', 'r') as f:
    all_shifts = json.load(f)

# Filter out unnamed shifts as requested
all_shifts = [s for s in all_shifts if s['name'].strip()]

print(f"Named shifts: {len(all_shifts)}")

# ==============================
# Color definitions from Master Daily Schedule PDF
# ==============================
# Section header fills (from the master schedule drawing colors)
FILL_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_EAST_LOT = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
FILL_SOUTH_LOT = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')  # Gray (0.749)
FILL_WEST_LOT = PatternFill(start_color='808080', end_color='808080', fill_type='solid')  # Darker gray (0.502)
FILL_CBC = PatternFill(start_color='C4D79B', end_color='C4D79B', fill_type='solid')  # Green (196,215,155)
FILL_BP_LOT = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')  # Blue (141,180,226)
FILL_BREAKER = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow
FILL_BLACK = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
FILL_LIGHT_BLUE = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

# Fonts
FONT_TITLE = Font(name='Arial', size=9, bold=True)
FONT_HEADER = Font(name='Arial', size=7, bold=True)
FONT_SECTION_TITLE = Font(name='Arial', size=10, bold=True)
FONT_DATA = Font(name='Arial', size=7)
FONT_SMALL = Font(name='Arial', size=5, bold=True)
FONT_DATA_BOLD = Font(name='Arial', size=7, bold=True)
FONT_WHITE_BOLD = Font(name='Arial', size=7, bold=True, color='FFFFFF')
FONT_WHITE_SMALL = Font(name='Arial', size=5, bold=True, color='FFFFFF')

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Day names in order (Sun first)
DAYS_ORDER = ['Sun', 'Mon', 'Tue', 'Wed', 'Thur', 'Fri', 'Sat']
DAY_FULL_NAMES = {
    'Sun': 'Sunday', 'Mon': 'Monday', 'Tue': 'Tuesday',
    'Wed': 'Wednesday', 'Thur': 'Thursday', 'Fri': 'Friday', 'Sat': 'Saturday'
}

LOT_ORDER = ['EAST LOT', 'SOUTH LOT', 'WEST LOT', 'CBC', 'BP LOT', 'BREAKER']
LOT_FILLS = {
    'EAST LOT': FILL_EAST_LOT,
    'SOUTH LOT': FILL_SOUTH_LOT,
    'WEST LOT': FILL_WEST_LOT,
    'CBC': FILL_CBC,
    'BP LOT': FILL_BP_LOT,
    'BREAKER': FILL_BREAKER,
}

def normalize_time(t):
    """Normalize time strings to HHMM format"""
    t = t.replace(':', '').strip()
    # Remove extra leading zeros or fix obvious typos
    t = re.sub(r'^0+(\d{4})$', r'\1', t)
    # Pad to at least 4 chars
    if len(t) == 3:
        t = '0' + t
    # Handle cases like "19000-0230" -> "1900"
    if len(t) > 4:
        t = t[:4]
    return t

def format_time_display(t):
    """Format HHMM to HH:MM for display"""
    t = normalize_time(t)
    if len(t) >= 4:
        return f"{t[:2]}:{t[2:4]}"
    return t

def parse_hours(hrs_str):
    """Parse hours string to float"""
    try:
        return float(hrs_str)
    except (ValueError, TypeError):
        return 0

def get_employees_for_day_lot(day_name, lot_name):
    """Get all employees working at a specific lot on a specific day, sorted by start time"""
    employees = []
    for s in all_shifts:
        d = s['days'].get(day_name, {})
        if d.get('lot') == lot_name and d.get('time') and d['time'] != 'OFF':
            start = d.get('start', '')
            end = d.get('end', '')
            hrs = d.get('hrs', '0')
            employees.append({
                'name': s['name'],
                'shift': s['shift'],
                'start': normalize_time(start) if start else '',
                'end': normalize_time(end) if end else '',
                'hrs': hrs,
                'time': d['time']
            })
    
    # Sort by start time
    employees.sort(key=lambda x: x['start'] if x['start'] else '9999')
    return employees

def write_section_header_cols(ws, row, start_col, label_text, lot_fill, is_right_section=False):
    """Write the column headers for a lot section: BUS#, Relief, [LOT NAME], Start Time, End Time, Hours, Trade Day x2, Overtime, Trade Day"""
    col = start_col
    
    # BUS# column
    cell = ws.cell(row=row, column=col, value='BUS#')
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Relief column
    cell = ws.cell(row=row, column=col, value='Relief')
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Lot name (merged across name cell)
    cell = ws.cell(row=row, column=col, value=label_text)
    cell.font = FONT_SECTION_TITLE
    cell.fill = lot_fill if lot_fill else FILL_WHITE
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Start Time
    cell = ws.cell(row=row, column=col, value='Start Time')
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # End Time
    cell = ws.cell(row=row, column=col, value='End Time')
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Hours
    cell = ws.cell(row=row, column=col, value='Hours')
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Trade Day
    cell = ws.cell(row=row, column=col, value='Trade Day')
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Trade Day
    cell = ws.cell(row=row, column=col, value='Trade Day')
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Overtime
    cell = ws.cell(row=row, column=col, value='Overtime')
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Trade Day
    cell = ws.cell(row=row, column=col, value='Trade Day')
    cell.font = FONT_SMALL
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    
    return col

def write_employee_row(ws, row, start_col, emp):
    """Write one employee data row"""
    col = start_col
    
    # BUS# (empty)
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER
    cell.font = FONT_DATA
    col += 1
    
    # Relief (empty)
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER
    cell.font = FONT_DATA
    col += 1
    
    # Name
    cell = ws.cell(row=row, column=col, value=emp['name'])
    cell.font = FONT_DATA
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    col += 1
    
    # Start Time
    cell = ws.cell(row=row, column=col, value=format_time_display(emp['start']) if emp['start'] else '')
    cell.font = FONT_DATA
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # End Time
    cell = ws.cell(row=row, column=col, value=format_time_display(emp['end']) if emp['end'] else '')
    cell.font = FONT_DATA
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Hours
    hrs_val = parse_hours(emp['hrs'])
    cell = ws.cell(row=row, column=col, value=hrs_val if hrs_val else '')
    cell.font = FONT_DATA
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Trade Day (empty)
    for _ in range(2):
        cell = ws.cell(row=row, column=col, value='')
        cell.border = THIN_BORDER
        cell.font = FONT_DATA
        col += 1
    
    # Overtime (empty)
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER
    cell.font = FONT_DATA
    col += 1
    
    # Trade Day (empty)
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER
    cell.font = FONT_DATA
    
    return col

def write_empty_row(ws, row, start_col):
    """Write an empty data row with borders and 0:00 in hours"""
    col = start_col
    for i in range(10):
        cell = ws.cell(row=row, column=col + i, value='')
        cell.border = THIN_BORDER
        cell.font = FONT_DATA
        cell.alignment = ALIGN_CENTER
    # Put 0:00 in the hours column
    ws.cell(row=row, column=start_col + 5, value='0:00').font = FONT_DATA

def create_daily_sheet(wb, day_name):
    """Create one daily schedule sheet"""
    ws = wb.create_sheet(title=DAY_FULL_NAMES[day_name])
    
    # ===== Page setup =====
    ws.sheet_properties.pageSetUpPr = None
    
    # Column widths - Left side (A-J) and Right side (K-T)
    # Left: BUS#(A), Relief(B), Name(C), StartTime(D), EndTime(E), Hours(F), TradeDay(G), TradeDay(H), Overtime(I), TradeDay(J)
    # Right: BUS#(K), Relief(L), Name(M), StartTime(N), EndTime(O), Hours(P), TradeDay(Q), TradeDay(R), Overtime(S), TradeDay(T)
    col_widths = {
        'A': 4, 'B': 5, 'C': 16, 'D': 8, 'E': 8, 'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5,
        'K': 4, 'L': 5, 'M': 16, 'N': 8, 'O': 8, 'P': 5, 'Q': 5, 'R': 5, 'S': 5, 'T': 5,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # ===== Header rows =====
    current_row = 1
    
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    cell = ws.cell(row=1, column=1, value='ABM Parking Services')
    cell.font = Font(name='Arial', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
    cell = ws.cell(row=2, column=1, value='Los Angeles Employee Shuttle')
    cell.font = Font(name='Arial', size=8)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=20)
    cell = ws.cell(row=3, column=1, value='Staffing Schedule')
    cell.font = Font(name='Arial', size=8)
    cell.alignment = Alignment(horizontal='center')
    
    current_row = 4
    
    # Shift / Day / Date row
    cell = ws.cell(row=current_row, column=1, value='Shift:')
    cell.font = FONT_DATA_BOLD
    cell = ws.cell(row=current_row, column=8, value='DAY:')
    cell.font = FONT_DATA_BOLD
    cell = ws.cell(row=current_row, column=9, value=DAY_FULL_NAMES[day_name])
    cell.font = FONT_DATA_BOLD
    cell = ws.cell(row=current_row, column=15, value='Date:')
    cell.font = FONT_DATA_BOLD
    
    current_row = 5
    
    # ===== MOD/Supervisor/Dispatcher Section =====
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    cell = ws.cell(row=current_row, column=1, value='MOD/Supervisor/Dispatcher')
    cell.font = FONT_SECTION_TITLE
    cell.border = THIN_BORDER
    for c in range(1, 11):
        ws.cell(row=current_row, column=c).border = THIN_BORDER
    
    # Column headers for MOD section (left side)
    current_row += 1
    col = 3  # Start at Name column
    for header in ['Start time', 'End Time', 'Hours', 'Trade Day', 'Trade Day', 'Overtime', 'Trade Day']:
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = FONT_SMALL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        col += 1
    # BUS# and Relief
    ws.cell(row=current_row, column=1).border = THIN_BORDER
    ws.cell(row=current_row, column=2).border = THIN_BORDER
    
    # Absences header on right side
    ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row, end_column=20)
    cell = ws.cell(row=current_row, column=11, value='Absences/ Lates/ NCNS/ FMLA/ LOA/ VAC/ Suspension')
    cell.font = Font(name='Arial', size=6, bold=True)
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    for c in range(11, 21):
        ws.cell(row=current_row, column=c).border = THIN_BORDER
    
    current_row += 1
    
    # MOD roles
    mod_roles = ['MOD', 'SUPERVISOR', 'SUPERVISOR', 'AMBASSADOR', 'DISPATCHER']
    for role in mod_roles:
        cell = ws.cell(row=current_row, column=1, value=role)
        cell.font = Font(name='Arial', size=6, bold=True)
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        for c in range(2, 11):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1
    
    # ===== MAIN BODY: Lot sections =====
    # Layout: Left side has 3 lots, Right side has 3 lots
    # Left: EAST LOT, WEST LOT, BREAKER
    # Right: SOUTH LOT, CBC, BP LOT
    
    # The Master Schedule PDF has a 2-column layout:
    # Left side (cols 1-10): EAST LOT, then WEST LOT, then BREAKER
    # Right side (cols 11-20): SOUTH LOT, then CBC, then BP LOT
    
    left_lots = ['EAST LOT', 'WEST LOT', 'BREAKER']
    right_lots = ['SOUTH LOT', 'CBC', 'BP LOT']
    
    # Get employee data for each lot
    lot_employees = {}
    for lot in LOT_ORDER:
        lot_employees[lot] = get_employees_for_day_lot(day_name, lot)
    
    # Calculate row counts to align left and right sections
    # Each lot section: 1 header row + N data rows
    left_data = []
    right_data = []
    for lot in left_lots:
        emps = lot_employees.get(lot, [])
        left_data.append((lot, emps))
    for lot in right_lots:
        emps = lot_employees.get(lot, [])
        right_data.append((lot, emps))
    
    # Write left and right sections - they share the same row space
    # Track the starting row for each section pair
    section_start_row = current_row
    
    # We need to interleave sections vertically
    # Left: EAST LOT header + data, then WEST LOT header + data, then BREAKER header + data
    # Right: SOUTH LOT header + data (starts at same row as EAST LOT data start), then CBC, then BP LOT
    
    # Let's calculate rows needed for each section
    left_sections = []
    for lot, emps in left_data:
        max_rows = max(len(emps), 2)  # At least 2 empty rows
        left_sections.append((lot, emps, max_rows))
    
    right_sections = []
    for lot, emps in right_data:
        max_rows = max(len(emps), 2)  # At least 2 empty rows
        right_sections.append((lot, emps, max_rows))
    
    # The master schedule PDF layout:
    # Row 1: EAST LOT header (left cols 1-10)
    # Rows 2..N: EAST LOT data (left) | from row where SOUTH LOT starts (right)
    # The SOUTH LOT header appears partway down the EAST LOT section
    # Then WEST LOT below EAST LOT on the left | CBC below SOUTH LOT on the right
    # Then BREAKER below WEST LOT | BP LOT below CBC
    
    # For simplicity, let's align sections side by side:
    # Write all left sections starting at section_start_row
    # Write all right sections starting at the same places
    
    left_row = section_start_row
    right_row = section_start_row
    
    # Track total hours
    total_hours = 0
    
    # ===== Write LEFT sections =====
    for lot, emps, max_rows in left_sections:
        # Section header
        write_section_header_cols(ws, left_row, 1, lot, LOT_FILLS.get(lot, FILL_WHITE))
        left_row += 1
        
        # Employee data rows
        for emp in emps:
            write_employee_row(ws, left_row, 1, emp)
            total_hours += parse_hours(emp['hrs'])
            left_row += 1
        
        # Fill remaining empty rows to minimum
        remaining = max_rows - len(emps)
        for _ in range(remaining):
            write_empty_row(ws, left_row, 1)
            left_row += 1
    
    # ===== Write RIGHT sections =====
    # Right sections start at the same row as the first left data row
    right_row = section_start_row
    
    # We need to align right sections to start after the EAST LOT header
    # Find where EAST LOT data starts and place SOUTH LOT header there
    east_emps_count = max(len(lot_employees.get('EAST LOT', [])), 2)
    
    for lot, emps, max_rows in right_sections:
        # Section header
        write_section_header_cols(ws, right_row, 11, lot, LOT_FILLS.get(lot, FILL_WHITE), True)
        right_row += 1
        
        # Employee data rows
        for emp in emps:
            write_employee_row(ws, right_row, 11, emp)
            total_hours += parse_hours(emp['hrs'])
            right_row += 1
        
        # Fill remaining empty rows
        remaining = max_rows - len(emps)
        for _ in range(remaining):
            write_empty_row(ws, right_row, 11)
            right_row += 1
    
    # ===== Footer section =====
    footer_row = max(left_row, right_row) + 1
    
    # Notes section
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=10)
    cell = ws.cell(row=footer_row, column=1, value='NOTES (CALL OFFS, INCIDENTS, OPERATIONAL CHANGES, DETOURS)')
    cell.font = FONT_WHITE_BOLD
    cell.fill = FILL_BLACK
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    for c in range(11, 21):
        ws.cell(row=footer_row, column=c).border = THIN_BORDER
    
    footer_row += 1
    # Empty notes rows
    for _ in range(3):
        for c in range(1, 21):
            ws.cell(row=footer_row, column=c).border = THIN_BORDER
        footer_row += 1
    
    # Hours summary row
    cell = ws.cell(row=footer_row, column=1, value='HOURS SCHEDULED:')
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    
    cell = ws.cell(row=footer_row, column=5, value=total_hours)
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    
    cell = ws.cell(row=footer_row, column=7, value='ADDITIONAL HOURS:')
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    
    cell = ws.cell(row=footer_row, column=12, value='TOTAL HOURS:')
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    
    cell = ws.cell(row=footer_row, column=15, value=total_hours)
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    
    footer_row += 1
    
    cell = ws.cell(row=footer_row, column=1, value='EXPECTED HOURS')
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    
    cell = ws.cell(row=footer_row, column=5, value=305)
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    
    # Apply borders to footer
    for c in range(1, 21):
        for r in range(footer_row - 1, footer_row + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
    
    # Print layout settings
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    return ws

# ===== Create workbook =====
wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

for day in DAYS_ORDER:
    print(f"Creating {DAY_FULL_NAMES[day]} sheet...")
    create_daily_sheet(wb, day)
    
    # Print summary for this day
    for lot in LOT_ORDER:
        emps = get_employees_for_day_lot(day, lot)
        total_hrs = sum(parse_hours(e['hrs']) for e in emps)
        print(f"  {lot}: {len(emps)} employees, {total_hrs} hrs")

# Save
output_path = 'Master Daily Schedule 2026.xlsx'
wb.save(output_path)
print(f"\nSaved to: {output_path}")
