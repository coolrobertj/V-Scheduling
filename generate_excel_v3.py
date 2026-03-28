import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load extracted bid data
with open('bid_data.json', 'r') as f:
    all_shifts = json.load(f)

# Filter out unnamed shifts as requested
all_shifts = [s for s in all_shifts if s['name'].strip()]

print(f"Named shifts: {len(all_shifts)}")

# ==============================
# Color definitions from Master Daily Schedule PDF
# ==============================
FILL_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_EAST_LOT = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_SOUTH_LOT = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
FILL_WEST_LOT = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
FILL_CBC = PatternFill(start_color='C4D79B', end_color='C4D79B', fill_type='solid')
FILL_BP_LOT = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
FILL_BREAKER = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
FILL_BLACK = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

# Fonts
FONT_SECTION_TITLE = Font(name='Arial', size=13, bold=True)
FONT_DATA = Font(name='Arial', size=10)
FONT_SMALL = Font(name='Arial', size=8, bold=True)
FONT_DATA_BOLD = Font(name='Arial', size=10, bold=True)
FONT_WHITE_BOLD = Font(name='Arial', size=10, bold=True, color='FFFFFF')

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=False)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=False)
ALIGN_ROTATED = Alignment(horizontal='center', vertical='center', wrap_text=False, text_rotation=90)

# Day names in order (Sun first)
DAYS_ORDER = ['Sun', 'Mon', 'Tue', 'Wed', 'Thur', 'Fri', 'Sat']
DAY_FULL_NAMES = {
    'Sun': 'Sunday', 'Mon': 'Monday', 'Tue': 'Tuesday',
    'Wed': 'Wednesday', 'Thur': 'Thursday', 'Fri': 'Friday', 'Sat': 'Saturday'
}

# Shift types
SHIFT_TYPES = ['Day', 'Swing', 'Graveyard']

LOT_ORDER = ['EAST LOT', 'SOUTH LOT', 'WEST LOT', 'CBC', 'BP LOT', 'BREAKER']
LOT_FILLS = {
    'EAST LOT': FILL_EAST_LOT,
    'SOUTH LOT': FILL_SOUTH_LOT,
    'WEST LOT': FILL_WEST_LOT,
    'CBC': FILL_CBC,
    'BP LOT': FILL_BP_LOT,
    'BREAKER': FILL_BREAKER,
}

def normalize_time(t):
    t = t.replace(':', '').strip()
    t = re.sub(r'^0+(\d{4})$', r'\1', t)
    if len(t) == 3:
        t = '0' + t
    if len(t) > 4:
        t = t[:4]
    return t

def format_time_display(t):
    t = normalize_time(t)
    if len(t) >= 4:
        return f"{t[:2]}:{t[2:4]}"
    return t

def time_to_excel_value(t):
    """Convert a normalized time string to a military time integer (e.g. 700 for 07:00)."""
    t = normalize_time(t)
    if len(t) >= 4:
        try:
            hh = int(t[:2])
            mm = int(t[2:4])
            return hh * 100 + mm
        except ValueError:
            return None
    return None

TIME_FORMAT = '00":"00'

def parse_hours(hrs_str):
    try:
        return float(hrs_str)
    except (ValueError, TypeError):
        return 0

def classify_shift_type(start_time_str):
    t = normalize_time(start_time_str)
    if len(t) < 4:
        return 'Day'
    try:
        hh = int(t[:2])
        mm = int(t[2:4])
        total_minutes = hh * 60 + mm
    except (ValueError, IndexError):
        return 'Day'
    if 0 <= total_minutes <= 10 * 60 + 59:
        return 'Day'
    elif 11 * 60 <= total_minutes <= 18 * 60 + 59:
        return 'Swing'
    else:
        return 'Graveyard'

def get_employees_for_day_lot_shift(day_name, lot_name, shift_type):
    employees = []
    for s in all_shifts:
        d = s['days'].get(day_name, {})
        if d.get('lot') == lot_name and d.get('time') and d['time'] != 'OFF':
            start = d.get('start', '')
            end = d.get('end', '')
            hrs = d.get('hrs', '0')
            emp_shift_type = classify_shift_type(start) if start else 'Day'
            if emp_shift_type == shift_type:
                employees.append({
                    'name': s['name'],
                    'shift': s['shift'],
                    'start': normalize_time(start) if start else '',
                    'end': normalize_time(end) if end else '',
                    'hrs': hrs,
                    'time': d['time']
                })
    employees.sort(key=lambda x: x['start'] if x['start'] else '9999')
    return employees

def get_employees_for_day_lot(day_name, lot_name):
    """Get ALL employees for a lot on a day (no shift filter) - for Master tab"""
    employees = []
    for s in all_shifts:
        d = s['days'].get(day_name, {})
        if d.get('lot') == lot_name and d.get('time') and d['time'] != 'OFF':
            start = d.get('start', '')
            end = d.get('end', '')
            hrs = d.get('hrs', '0')
            employees.append({
                'name': s['name'],
                'shift': s['shift'],
                'start': normalize_time(start) if start else '',
                'end': normalize_time(end) if end else '',
                'hrs': hrs,
                'time': d['time']
            })
    employees.sort(key=lambda x: x['start'] if x['start'] else '9999')
    return employees

def set_col_widths(ws):
    col_widths = {
        'A': 10.36, 'B': 5, 'C': 16, 'D': 8, 'E': 8, 'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5,
        'K': 10.36, 'L': 5, 'M': 16, 'N': 8, 'O': 8, 'P': 5, 'Q': 5, 'R': 5, 'S': 5, 'T': 5,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

def write_section_header_cols(ws, row, start_col, label_text, lot_fill):
    col = start_col
    for val, font in [('BUS#', FONT_SMALL), ('Relief', FONT_SMALL)]:
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        col += 1
    # Lot name
    cell = ws.cell(row=row, column=col, value=label_text)
    cell.font = FONT_SECTION_TITLE
    cell.fill = lot_fill if lot_fill else FILL_WHITE
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    for val in ['Start Time', 'End Time', 'Hours', 'Trade Day', 'Trade Day', 'Overtime', 'Trade Day']:
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = FONT_SMALL
        if val in ('Trade Day', 'Overtime'):
            cell.alignment = ALIGN_ROTATED
        else:
            cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        col += 1
    return col

def write_employee_row(ws, row, start_col, emp):
    col = start_col
    # BUS#
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER; cell.font = FONT_DATA
    col += 1
    # Relief
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER; cell.font = FONT_DATA
    col += 1
    # Name
    cell = ws.cell(row=row, column=col, value=emp['name'])
    cell.font = FONT_DATA; cell.alignment = ALIGN_LEFT; cell.border = THIN_BORDER
    col += 1
    # Start Time
    time_val = time_to_excel_value(emp['start']) if emp['start'] else None
    cell = ws.cell(row=row, column=col, value=time_val if time_val is not None else '')
    cell.font = FONT_DATA; cell.alignment = ALIGN_CENTER; cell.border = THIN_BORDER
    cell.number_format = TIME_FORMAT
    col += 1
    # End Time
    time_val = time_to_excel_value(emp['end']) if emp['end'] else None
    cell = ws.cell(row=row, column=col, value=time_val if time_val is not None else '')
    cell.font = FONT_DATA; cell.alignment = ALIGN_CENTER; cell.border = THIN_BORDER
    cell.number_format = TIME_FORMAT
    col += 1
    # Hours (formula: calculated from Start and End)
    sc = get_column_letter(start_col + 3)  # Start col letter
    ec = get_column_letter(start_col + 4)  # End col letter
    formula = (f'=IF(AND({sc}{row}<>"",{ec}{row}<>""),'
               f'INT(IF({ec}{row}<{sc}{row},'
               f'24+(INT({ec}{row}/100)+MOD({ec}{row},100)/60)-(INT({sc}{row}/100)+MOD({sc}{row},100)/60),'
               f'(INT({ec}{row}/100)+MOD({ec}{row},100)/60)-(INT({sc}{row}/100)+MOD({sc}{row},100)/60))),"")')
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = FONT_DATA; cell.alignment = ALIGN_CENTER; cell.border = THIN_BORDER
    cell.number_format = '0'
    col += 1
    # Trade Day, Trade Day, Overtime, Trade Day
    for _ in range(4):
        cell = ws.cell(row=row, column=col, value='')
        cell.border = THIN_BORDER; cell.font = FONT_DATA
        col += 1
    return col

def write_empty_row(ws, row, start_col):
    col = start_col
    # BUS#
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER; cell.font = FONT_DATA
    col += 1
    # Relief
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER; cell.font = FONT_DATA
    col += 1
    # Name
    cell = ws.cell(row=row, column=col, value='')
    cell.font = FONT_DATA; cell.alignment = ALIGN_LEFT; cell.border = THIN_BORDER
    col += 1
    # Start Time
    cell = ws.cell(row=row, column=col, value='')
    cell.font = FONT_DATA; cell.alignment = ALIGN_CENTER; cell.border = THIN_BORDER
    cell.number_format = TIME_FORMAT
    col += 1
    # End Time
    cell = ws.cell(row=row, column=col, value='')
    cell.font = FONT_DATA; cell.alignment = ALIGN_CENTER; cell.border = THIN_BORDER
    cell.number_format = TIME_FORMAT
    col += 1
    # Hours (formula: calculated from Start and End)
    sc = get_column_letter(start_col + 3)  # Start col letter
    ec = get_column_letter(start_col + 4)  # End col letter
    formula = (f'=IF(AND({sc}{row}<>"",{ec}{row}<>""),'
               f'INT(IF({ec}{row}<{sc}{row},'
               f'24+(INT({ec}{row}/100)+MOD({ec}{row},100)/60)-(INT({sc}{row}/100)+MOD({sc}{row},100)/60),'
               f'(INT({ec}{row}/100)+MOD({ec}{row},100)/60)-(INT({sc}{row}/100)+MOD({sc}{row},100)/60))),"")')
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = FONT_DATA; cell.alignment = ALIGN_CENTER; cell.border = THIN_BORDER
    cell.number_format = '0'
    col += 1
    # Trade Day, Trade Day, Overtime, Trade Day
    for _ in range(4):
        cell = ws.cell(row=row, column=col, value='')
        cell.border = THIN_BORDER; cell.font = FONT_DATA
        col += 1

def write_absences_section(ws, header_row, end_row):
    """Write the 2-column absences section on the right side (cols 11-20) from header_row+1 to end_row.
    Two columns: col 11-15 = Name1, col 16-20 = Name2 (each with borders)."""
    for r in range(header_row + 1, end_row + 1):
        for c in range(11, 21):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            cell.font = FONT_DATA

def write_header_and_mod(ws, shift_label, day_label):
    """Write the common header, shift/day/date, and MOD section. Returns the current_row after MOD."""
    # 3 blank rows at top
    start = 4
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=20)
    cell = ws.cell(row=start, column=1, value='ABM Parking Services')
    cell.font = Font(name='Arial', size=12, bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=start+1, start_column=1, end_row=start+1, end_column=20)
    cell = ws.cell(row=start+1, column=1, value='Los Angeles Employee Shuttle')
    cell.font = Font(name='Arial', size=11)
    cell.alignment = Alignment(horizontal='center')

    ws.merge_cells(start_row=start+2, start_column=1, end_row=start+2, end_column=20)
    cell = ws.cell(row=start+2, column=1, value='Staffing Schedule')
    cell.font = Font(name='Arial', size=11)
    cell.alignment = Alignment(horizontal='center')

    current_row = start + 3
    # Shift / Day / Date
    ws.cell(row=current_row, column=1, value='Shift:').font = FONT_DATA_BOLD
    ws.cell(row=current_row, column=2, value=shift_label).font = FONT_DATA_BOLD
    ws.cell(row=current_row, column=8, value='DAY:').font = FONT_DATA_BOLD
    ws.cell(row=current_row, column=9, value=day_label).font = FONT_DATA_BOLD
    ws.cell(row=current_row, column=15, value='Date:').font = FONT_DATA_BOLD

    current_row += 1
    # MOD/Supervisor/Dispatcher header
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    cell = ws.cell(row=current_row, column=1, value='MOD/Supervisor/Dispatcher')
    cell.font = FONT_SECTION_TITLE; cell.border = THIN_BORDER
    for c in range(1, 11):
        ws.cell(row=current_row, column=c).border = THIN_BORDER

    current_row += 1
    # Column headers for MOD
    col = 3
    for header in ['Start time', 'End Time', 'Hours', 'Trade Day', 'Trade Day', 'Overtime', 'Trade Day']:
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = FONT_SMALL
        if header in ('Trade Day', 'Overtime'):
            cell.alignment = ALIGN_ROTATED
        else:
            cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        col += 1
    ws.cell(row=current_row, column=1).border = THIN_BORDER
    ws.cell(row=current_row, column=2).border = THIN_BORDER

    # Absences header on right side - 2 columns
    absences_header_row = current_row
    ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row, end_column=20)
    cell = ws.cell(row=current_row, column=11, value='Absences/ Lates/ NCNS/ FMLA/ LOA/ VAC/ Suspension')
    cell.font = Font(name='Arial', size=9, bold=True)
    cell.alignment = ALIGN_CENTER; cell.border = THIN_BORDER
    for c in range(11, 21):
        ws.cell(row=current_row, column=c).border = THIN_BORDER

    current_row += 1
    # Absences sub-headers: two name columns
    ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row, end_column=15)
    cell = ws.cell(row=current_row, column=11, value='Name')
    cell.font = FONT_SMALL; cell.alignment = ALIGN_CENTER; cell.border = THIN_BORDER
    for c in range(11, 16):
        ws.cell(row=current_row, column=c).border = THIN_BORDER

    ws.merge_cells(start_row=current_row, start_column=16, end_row=current_row, end_column=20)
    cell = ws.cell(row=current_row, column=16, value='Name')
    cell.font = FONT_SMALL; cell.alignment = ALIGN_CENTER; cell.border = THIN_BORDER
    for c in range(16, 21):
        ws.cell(row=current_row, column=c).border = THIN_BORDER

    # MOD roles (starting at same row as absences sub-header)
    mod_roles = ['MOD', 'SUPERVISOR', 'SUPERVISOR', 'AMBASSADOR', 'DISPATCHER']
    mod_start_row = current_row
    for role in mod_roles:
        cell = ws.cell(row=current_row, column=1, value=role)
        cell.font = Font(name='Arial', size=9, bold=True)
        cell.alignment = ALIGN_LEFT; cell.border = THIN_BORDER
        for c in range(2, 11):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        # Absences rows (2 merged columns for name entries)
        ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row, end_column=15)
        ws.cell(row=current_row, column=11).border = THIN_BORDER
        for c in range(11, 16):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        ws.merge_cells(start_row=current_row, start_column=16, end_row=current_row, end_column=20)
        ws.cell(row=current_row, column=16).border = THIN_BORDER
        for c in range(16, 21):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1

    return current_row, absences_header_row

def write_lot_sections(ws, current_row, lot_employees_dict, left_lots, right_lots):
    """Write left and right lot sections. Returns (left_row, right_row, total_hours)."""
    total_hours = 0

    # ===== Write LEFT sections =====
    left_row = current_row
    left_section_rows = {}  # Track start/end for each left section
    for lot in left_lots:
        emps = lot_employees_dict.get(lot, [])
        max_rows = max(len(emps), 2)
        write_section_header_cols(ws, left_row, 1, lot, LOT_FILLS.get(lot, FILL_WHITE))
        section_start = left_row + 1
        left_row += 1
        for emp in emps:
            write_employee_row(ws, left_row, 1, emp)
            total_hours += parse_hours(emp['hrs'])
            left_row += 1
        remaining = max_rows - len(emps)
        for _ in range(remaining):
            write_empty_row(ws, left_row, 1)
            left_row += 1
        left_section_rows[lot] = (section_start, left_row - 1)

    # ===== Write RIGHT sections =====
    right_row = current_row
    right_section_rows = {}
    for lot in right_lots:
        emps = lot_employees_dict.get(lot, [])
        max_rows = max(len(emps), 2)
        write_section_header_cols(ws, right_row, 11, lot, LOT_FILLS.get(lot, FILL_WHITE))
        section_start = right_row + 1
        right_row += 1
        for emp in emps:
            write_employee_row(ws, right_row, 11, emp)
            total_hours += parse_hours(emp['hrs'])
            right_row += 1
        remaining = max_rows - len(emps)
        for _ in range(remaining):
            write_empty_row(ws, right_row, 11)
            right_row += 1
        right_section_rows[lot] = (section_start, right_row - 1)

    # ===== FIX: Extend right-side borders down to match left side =====
    # Use write_empty_row so formats (time, number) carry over to blank rows
    final_row = max(left_row, right_row)
    if right_row < final_row:
        for r in range(right_row, final_row):
            write_empty_row(ws, r, 11)
    if left_row < final_row:
        for r in range(left_row, final_row):
            write_empty_row(ws, r, 1)

    return final_row, total_hours

def write_footer(ws, footer_start, total_hours):
    """Write notes and hours footer section."""
    footer_row = footer_start

    # Notes section
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=10)
    cell = ws.cell(row=footer_row, column=1, value='NOTES (CALL OFFS, INCIDENTS, OPERATIONAL CHANGES, DETOURS)')
    cell.font = FONT_WHITE_BOLD; cell.fill = FILL_BLACK; cell.alignment = ALIGN_LEFT; cell.border = THIN_BORDER
    for c in range(11, 21):
        ws.cell(row=footer_row, column=c).border = THIN_BORDER

    footer_row += 1
    for _ in range(3):
        for c in range(1, 21):
            ws.cell(row=footer_row, column=c).border = THIN_BORDER
        footer_row += 1

    # Hours summary
    ws.cell(row=footer_row, column=1, value='HOURS SCHEDULED:').font = FONT_DATA_BOLD
    ws.cell(row=footer_row, column=1).border = THIN_BORDER
    cell = ws.cell(row=footer_row, column=5, value=total_hours)
    cell.font = FONT_DATA_BOLD; cell.border = THIN_BORDER; cell.alignment = ALIGN_CENTER
    ws.cell(row=footer_row, column=7, value='ADDITIONAL HOURS:').font = FONT_DATA_BOLD
    ws.cell(row=footer_row, column=7).border = THIN_BORDER
    ws.cell(row=footer_row, column=12, value='TOTAL HOURS:').font = FONT_DATA_BOLD
    ws.cell(row=footer_row, column=12).border = THIN_BORDER
    cell = ws.cell(row=footer_row, column=15, value=total_hours)
    cell.font = FONT_DATA_BOLD; cell.border = THIN_BORDER; cell.alignment = ALIGN_CENTER

    footer_row += 1
    ws.cell(row=footer_row, column=1, value='EXPECTED HOURS').font = FONT_DATA_BOLD
    ws.cell(row=footer_row, column=1).border = THIN_BORDER
    cell = ws.cell(row=footer_row, column=5, value=305)
    cell.font = FONT_DATA_BOLD; cell.border = THIN_BORDER; cell.alignment = ALIGN_CENTER

    for c in range(1, 21):
        for r in range(footer_row - 1, footer_row + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Page setup - fit entire sheet on a single page with minimal margins
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.1
    ws.page_margins.right = 0.1
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1
    last_col = get_column_letter(20)
    ws.print_area = f'A1:{last_col}{footer_row}'

# ===========================
# CREATE MASTER (blank) TAB
# ===========================
def create_master_sheet(wb):
    """Create the blank Master template as the first tab, matching the PDF exactly."""
    ws = wb.create_sheet(title='Master', index=0)
    set_col_widths(ws)

    current_row, abs_header_row = write_header_and_mod(ws, '', '', )

    left_lots = ['EAST LOT', 'WEST LOT', 'BREAKER']
    right_lots = ['SOUTH LOT', 'CBC', 'BP LOT']

    # Empty lot sections with fixed row counts matching the PDF template
    # PDF has: EAST LOT ~12 rows, SOUTH LOT ~12 rows, WEST LOT ~7 rows, CBC ~7 rows,
    # BREAKER ~8 rows, BP LOT ~1 row
    template_rows = {
        'EAST LOT': 12, 'SOUTH LOT': 12,
        'WEST LOT': 7, 'CBC': 7,
        'BREAKER': 8, 'BP LOT': 2,
    }

    lot_employees_dict = {lot: [] for lot in LOT_ORDER}

    # Write LEFT sections
    left_row = current_row
    for lot in left_lots:
        num_rows = template_rows.get(lot, 5)
        write_section_header_cols(ws, left_row, 1, lot, LOT_FILLS.get(lot, FILL_WHITE))
        left_row += 1
        for _ in range(num_rows):
            write_empty_row(ws, left_row, 1)
            left_row += 1

    # Write RIGHT sections
    right_row = current_row
    for lot in right_lots:
        num_rows = template_rows.get(lot, 5)
        write_section_header_cols(ws, right_row, 11, lot, LOT_FILLS.get(lot, FILL_WHITE))
        right_row += 1
        for _ in range(num_rows):
            write_empty_row(ws, right_row, 11)
            right_row += 1

    # Extend borders to align both sides
    final_row = max(left_row, right_row)
    if right_row < final_row:
        for r in range(right_row, final_row):
            write_empty_row(ws, r, 11)
    if left_row < final_row:
        for r in range(left_row, final_row):
            write_empty_row(ws, r, 1)

    write_footer(ws, final_row + 1, 0)
    return ws

# ===========================
# CREATE DAILY SHIFT TABS
# ===========================
def create_daily_shift_sheet(wb, day_name, shift_type):
    tab_name = f"{DAY_FULL_NAMES[day_name]} {shift_type}"
    if len(tab_name) > 31:
        tab_name = tab_name[:31]
    ws = wb.create_sheet(title=tab_name)
    set_col_widths(ws)

    current_row, abs_header_row = write_header_and_mod(ws, shift_type, DAY_FULL_NAMES[day_name])

    left_lots = ['EAST LOT', 'WEST LOT', 'BREAKER']
    right_lots = ['SOUTH LOT', 'CBC', 'BP LOT']

    lot_employees_dict = {}
    for lot in LOT_ORDER:
        lot_employees_dict[lot] = get_employees_for_day_lot_shift(day_name, lot, shift_type)

    final_row, total_hours = write_lot_sections(ws, current_row, lot_employees_dict, left_lots, right_lots)

    # Write absences rows alongside the lot data (from after MOD section to end of lots)
    # The absences column area already has borders from lot sections on right side.
    # We just need to ensure the absences area exists for the MOD rows (done in write_header_and_mod).

    write_footer(ws, final_row + 1, total_hours)
    return ws

# ===== Create workbook =====
wb = Workbook()
wb.remove(wb.active)

# 1) Master blank template tab
print("Creating: Master (blank template)")
create_master_sheet(wb)

# 2) 21 daily shift tabs
for day in DAYS_ORDER:
    for shift_type in SHIFT_TYPES:
        tab_label = f"{DAY_FULL_NAMES[day]} {shift_type}"
        print(f"Creating: {tab_label}")
        create_daily_shift_sheet(wb, day, shift_type)

        total_emps = 0
        total_hrs = 0
        for lot in LOT_ORDER:
            emps = get_employees_for_day_lot_shift(day, lot, shift_type)
            hrs = sum(parse_hours(e['hrs']) for e in emps)
            total_emps += len(emps)
            total_hrs += hrs
            if emps:
                print(f"  {lot}: {len(emps)} employees, {hrs} hrs")
        print(f"  TOTAL: {total_emps} employees, {total_hrs} hrs")

# Save
output_path = 'Master Daily Schedule 2026.xlsx'
wb.save(output_path)
print(f"\nSaved to: {output_path}")
