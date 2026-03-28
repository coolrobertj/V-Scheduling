import json
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load extracted bid data
with open('bid_data.json', 'r') as f:
    all_shifts = json.load(f)

# Filter out unnamed shifts as requested
all_shifts = [s for s in all_shifts if s['name'].strip()]

print(f"Named shifts: {len(all_shifts)}")

# ==============================
# Color definitions from Master Daily Schedule PDF
# ==============================
FILL_WHITE = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_EAST_LOT = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
FILL_SOUTH_LOT = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')
FILL_WEST_LOT = PatternFill(start_color='808080', end_color='808080', fill_type='solid')
FILL_CBC = PatternFill(start_color='C4D79B', end_color='C4D79B', fill_type='solid')
FILL_BP_LOT = PatternFill(start_color='8DB4E2', end_color='8DB4E2', fill_type='solid')
FILL_BREAKER = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
FILL_BLACK = PatternFill(start_color='000000', end_color='000000', fill_type='solid')

# Fonts
FONT_SECTION_TITLE = Font(name='Arial', size=10, bold=True)
FONT_DATA = Font(name='Arial', size=7)
FONT_SMALL = Font(name='Arial', size=5, bold=True)
FONT_DATA_BOLD = Font(name='Arial', size=7, bold=True)
FONT_WHITE_BOLD = Font(name='Arial', size=7, bold=True, color='FFFFFF')

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alignments
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Day names in order (Sun first)
DAYS_ORDER = ['Sun', 'Mon', 'Tue', 'Wed', 'Thur', 'Fri', 'Sat']
DAY_FULL_NAMES = {
    'Sun': 'Sunday', 'Mon': 'Monday', 'Tue': 'Tuesday',
    'Wed': 'Wednesday', 'Thur': 'Thursday', 'Fri': 'Friday', 'Sat': 'Saturday'
}

# Shift types
SHIFT_TYPES = ['Day', 'Swing', 'Graveyard']

LOT_ORDER = ['EAST LOT', 'SOUTH LOT', 'WEST LOT', 'CBC', 'BP LOT', 'BREAKER']
LOT_FILLS = {
    'EAST LOT': FILL_EAST_LOT,
    'SOUTH LOT': FILL_SOUTH_LOT,
    'WEST LOT': FILL_WEST_LOT,
    'CBC': FILL_CBC,
    'BP LOT': FILL_BP_LOT,
    'BREAKER': FILL_BREAKER,
}

def normalize_time(t):
    """Normalize time strings to HHMM format"""
    t = t.replace(':', '').strip()
    t = re.sub(r'^0+(\d{4})$', r'\1', t)
    if len(t) == 3:
        t = '0' + t
    if len(t) > 4:
        t = t[:4]
    return t

def format_time_display(t):
    """Format HHMM to HH:MM for display"""
    t = normalize_time(t)
    if len(t) >= 4:
        return f"{t[:2]}:{t[2:4]}"
    return t

def parse_hours(hrs_str):
    """Parse hours string to float"""
    try:
        return float(hrs_str)
    except (ValueError, TypeError):
        return 0

def classify_shift_type(start_time_str):
    """
    Classify shift type based on start time:
    - Day shift: 00:00 to 10:00 (0000-1059)
    - Swing shift: 11:00 to 18:00 (1100-1859)
    - Graveyard shift: 19:00 to 06:30 (1900-0630, wraps midnight)
    
    Since graveyard overlaps with day shift hours (00:00-06:30),
    we disambiguate: start times 00:00-06:30 are treated as Day shift
    per the ranges given (Day: 00:00-10:00 takes priority for these hours).
    Graveyard only applies to start times 19:00-23:59.
    """
    t = normalize_time(start_time_str)
    if len(t) < 4:
        return 'Day'  # default
    
    try:
        hh = int(t[:2])
        mm = int(t[2:4])
        total_minutes = hh * 60 + mm
    except (ValueError, IndexError):
        return 'Day'
    
    # Day shift: 00:00 (0) to 10:59 (659)
    if 0 <= total_minutes <= 10 * 60 + 59:
        return 'Day'
    # Swing shift: 11:00 (660) to 18:59 (1139)
    elif 11 * 60 <= total_minutes <= 18 * 60 + 59:
        return 'Swing'
    # Graveyard shift: 19:00 (1140) to 23:59 (1439)
    else:
        return 'Graveyard'

def get_employees_for_day_lot_shift(day_name, lot_name, shift_type):
    """Get employees for a specific lot, day, and shift type, sorted by start time"""
    employees = []
    for s in all_shifts:
        d = s['days'].get(day_name, {})
        if d.get('lot') == lot_name and d.get('time') and d['time'] != 'OFF':
            start = d.get('start', '')
            end = d.get('end', '')
            hrs = d.get('hrs', '0')
            
            # Classify this employee's shift type
            emp_shift_type = classify_shift_type(start) if start else 'Day'
            
            if emp_shift_type == shift_type:
                employees.append({
                    'name': s['name'],
                    'shift': s['shift'],
                    'start': normalize_time(start) if start else '',
                    'end': normalize_time(end) if end else '',
                    'hrs': hrs,
                    'time': d['time']
                })
    
    employees.sort(key=lambda x: x['start'] if x['start'] else '9999')
    return employees

def write_section_header_cols(ws, row, start_col, label_text, lot_fill):
    """Write the column headers for a lot section"""
    col = start_col
    
    headers_before = [('BUS#', FONT_SMALL), ('Relief', FONT_SMALL)]
    for val, font in headers_before:
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        col += 1
    
    # Lot name
    cell = ws.cell(row=row, column=col, value=label_text)
    cell.font = FONT_SECTION_TITLE
    cell.fill = lot_fill if lot_fill else FILL_WHITE
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    headers_after = ['Start Time', 'End Time', 'Hours', 'Trade Day', 'Trade Day', 'Overtime', 'Trade Day']
    for val in headers_after:
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = FONT_SMALL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        col += 1
    
    return col

def write_employee_row(ws, row, start_col, emp):
    """Write one employee data row"""
    col = start_col
    
    # BUS#
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER
    cell.font = FONT_DATA
    col += 1
    
    # Relief
    cell = ws.cell(row=row, column=col, value='')
    cell.border = THIN_BORDER
    cell.font = FONT_DATA
    col += 1
    
    # Name
    cell = ws.cell(row=row, column=col, value=emp['name'])
    cell.font = FONT_DATA
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    col += 1
    
    # Start Time
    cell = ws.cell(row=row, column=col, value=format_time_display(emp['start']) if emp['start'] else '')
    cell.font = FONT_DATA
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # End Time
    cell = ws.cell(row=row, column=col, value=format_time_display(emp['end']) if emp['end'] else '')
    cell.font = FONT_DATA
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Hours
    hrs_val = parse_hours(emp['hrs'])
    cell = ws.cell(row=row, column=col, value=hrs_val if hrs_val else '')
    cell.font = FONT_DATA
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    col += 1
    
    # Trade Day, Trade Day, Overtime, Trade Day (empty)
    for _ in range(4):
        cell = ws.cell(row=row, column=col, value='')
        cell.border = THIN_BORDER
        cell.font = FONT_DATA
        col += 1
    
    return col

def write_empty_row(ws, row, start_col):
    """Write an empty data row with borders"""
    for i in range(10):
        cell = ws.cell(row=row, column=start_col + i, value='')
        cell.border = THIN_BORDER
        cell.font = FONT_DATA
        cell.alignment = ALIGN_CENTER

def create_daily_shift_sheet(wb, day_name, shift_type):
    """Create one daily schedule sheet for a specific shift type"""
    tab_name = f"{DAY_FULL_NAMES[day_name]} {shift_type}"
    # Excel tab names max 31 chars
    if len(tab_name) > 31:
        tab_name = tab_name[:31]
    
    ws = wb.create_sheet(title=tab_name)
    
    # Column widths
    col_widths = {
        'A': 4, 'B': 5, 'C': 16, 'D': 8, 'E': 8, 'F': 5, 'G': 5, 'H': 5, 'I': 5, 'J': 5,
        'K': 4, 'L': 5, 'M': 16, 'N': 8, 'O': 8, 'P': 5, 'Q': 5, 'R': 5, 'S': 5, 'T': 5,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # ===== Header rows =====
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    cell = ws.cell(row=1, column=1, value='ABM Parking Services')
    cell.font = Font(name='Arial', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
    cell = ws.cell(row=2, column=1, value='Los Angeles Employee Shuttle')
    cell.font = Font(name='Arial', size=8)
    cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=20)
    cell = ws.cell(row=3, column=1, value='Staffing Schedule')
    cell.font = Font(name='Arial', size=8)
    cell.alignment = Alignment(horizontal='center')
    
    current_row = 4
    
    # Shift / Day / Date row - NOW FILLED IN
    cell = ws.cell(row=current_row, column=1, value='Shift:')
    cell.font = FONT_DATA_BOLD
    cell = ws.cell(row=current_row, column=2, value=shift_type)
    cell.font = FONT_DATA_BOLD
    
    cell = ws.cell(row=current_row, column=8, value='DAY:')
    cell.font = FONT_DATA_BOLD
    cell = ws.cell(row=current_row, column=9, value=DAY_FULL_NAMES[day_name])
    cell.font = FONT_DATA_BOLD
    
    cell = ws.cell(row=current_row, column=15, value='Date:')
    cell.font = FONT_DATA_BOLD
    
    current_row = 5
    
    # ===== MOD/Supervisor/Dispatcher Section =====
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    cell = ws.cell(row=current_row, column=1, value='MOD/Supervisor/Dispatcher')
    cell.font = FONT_SECTION_TITLE
    cell.border = THIN_BORDER
    for c in range(1, 11):
        ws.cell(row=current_row, column=c).border = THIN_BORDER
    
    current_row += 1
    col = 3
    for header in ['Start time', 'End Time', 'Hours', 'Trade Day', 'Trade Day', 'Overtime', 'Trade Day']:
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = FONT_SMALL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        col += 1
    ws.cell(row=current_row, column=1).border = THIN_BORDER
    ws.cell(row=current_row, column=2).border = THIN_BORDER
    
    # Absences header on right side
    ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row, end_column=20)
    cell = ws.cell(row=current_row, column=11, value='Absences/ Lates/ NCNS/ FMLA/ LOA/ VAC/ Suspension')
    cell.font = Font(name='Arial', size=6, bold=True)
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    for c in range(11, 21):
        ws.cell(row=current_row, column=c).border = THIN_BORDER
    
    current_row += 1
    
    # MOD roles
    mod_roles = ['MOD', 'SUPERVISOR', 'SUPERVISOR', 'AMBASSADOR', 'DISPATCHER']
    for role in mod_roles:
        cell = ws.cell(row=current_row, column=1, value=role)
        cell.font = Font(name='Arial', size=6, bold=True)
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        for c in range(2, 11):
            ws.cell(row=current_row, column=c).border = THIN_BORDER
        current_row += 1
    
    # ===== MAIN BODY: Lot sections =====
    left_lots = ['EAST LOT', 'WEST LOT', 'BREAKER']
    right_lots = ['SOUTH LOT', 'CBC', 'BP LOT']
    
    # Get employee data filtered by shift type
    lot_employees = {}
    for lot in LOT_ORDER:
        lot_employees[lot] = get_employees_for_day_lot_shift(day_name, lot, shift_type)
    
    left_data = [(lot, lot_employees.get(lot, [])) for lot in left_lots]
    right_data = [(lot, lot_employees.get(lot, [])) for lot in right_lots]
    
    # Track total hours
    total_hours = 0
    
    # ===== Write LEFT sections =====
    left_row = current_row
    for lot, emps in left_data:
        max_rows = max(len(emps), 2)
        write_section_header_cols(ws, left_row, 1, lot, LOT_FILLS.get(lot, FILL_WHITE))
        left_row += 1
        
        for emp in emps:
            write_employee_row(ws, left_row, 1, emp)
            total_hours += parse_hours(emp['hrs'])
            left_row += 1
        
        remaining = max_rows - len(emps)
        for _ in range(remaining):
            write_empty_row(ws, left_row, 1)
            left_row += 1
    
    # ===== Write RIGHT sections =====
    right_row = current_row
    for lot, emps in right_data:
        max_rows = max(len(emps), 2)
        write_section_header_cols(ws, right_row, 11, lot, LOT_FILLS.get(lot, FILL_WHITE))
        right_row += 1
        
        for emp in emps:
            write_employee_row(ws, right_row, 11, emp)
            total_hours += parse_hours(emp['hrs'])
            right_row += 1
        
        remaining = max_rows - len(emps)
        for _ in range(remaining):
            write_empty_row(ws, right_row, 11)
            right_row += 1
    
    # ===== Footer section =====
    footer_row = max(left_row, right_row) + 1
    
    # Notes section
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=10)
    cell = ws.cell(row=footer_row, column=1, value='NOTES (CALL OFFS, INCIDENTS, OPERATIONAL CHANGES, DETOURS)')
    cell.font = FONT_WHITE_BOLD
    cell.fill = FILL_BLACK
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    for c in range(11, 21):
        ws.cell(row=footer_row, column=c).border = THIN_BORDER
    
    footer_row += 1
    for _ in range(3):
        for c in range(1, 21):
            ws.cell(row=footer_row, column=c).border = THIN_BORDER
        footer_row += 1
    
    # Hours summary
    cell = ws.cell(row=footer_row, column=1, value='HOURS SCHEDULED:')
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    
    cell = ws.cell(row=footer_row, column=5, value=total_hours)
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    
    cell = ws.cell(row=footer_row, column=7, value='ADDITIONAL HOURS:')
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    
    cell = ws.cell(row=footer_row, column=12, value='TOTAL HOURS:')
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    
    cell = ws.cell(row=footer_row, column=15, value=total_hours)
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    
    footer_row += 1
    
    cell = ws.cell(row=footer_row, column=1, value='EXPECTED HOURS')
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    
    cell = ws.cell(row=footer_row, column=5, value=305)
    cell.font = FONT_DATA_BOLD
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    
    # Apply borders to footer
    for c in range(1, 21):
        for r in range(footer_row - 1, footer_row + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
    
    # Print layout settings
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    
    return ws

# ===== Create workbook with 21 tabs =====
wb = Workbook()
wb.remove(wb.active)

for day in DAYS_ORDER:
    for shift_type in SHIFT_TYPES:
        tab_label = f"{DAY_FULL_NAMES[day]} {shift_type}"
        print(f"Creating: {tab_label}")
        create_daily_shift_sheet(wb, day, shift_type)
        
        # Print summary
        total_emps = 0
        total_hrs = 0
        for lot in LOT_ORDER:
            emps = get_employees_for_day_lot_shift(day, lot, shift_type)
            hrs = sum(parse_hours(e['hrs']) for e in emps)
            total_emps += len(emps)
            total_hrs += hrs
            if emps:
                print(f"  {lot}: {len(emps)} employees, {hrs} hrs")
        print(f"  TOTAL: {total_emps} employees, {total_hrs} hrs")

# Save
output_path = 'Master Daily Schedule 2026.xlsx'
wb.save(output_path)
print(f"\nSaved to: {output_path}")
