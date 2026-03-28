import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

with open('bid_data.json') as f:
    data = json.load(f)

data = [s for s in data if s['name'].strip()]
data.sort(key=lambda x: x['name'])

wb = Workbook()
ws = wb.active
ws.title = 'Employee Schedule'

HEADER_FONT = Font(name='Arial', size=10, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
DATA_FONT = Font(name='Arial', size=10)
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

headers = ['Name', 'Shift #', 'Wkly Hrs', 'Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = BORDER

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 10
for letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws.column_dimensions[letter].width = 8

OFF_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
NONE_FILL = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

for row_idx, s in enumerate(data, 2):
    ws.cell(row=row_idx, column=1, value=s['name']).font = DATA_FONT
    ws.cell(row=row_idx, column=1).alignment = LEFT
    ws.cell(row=row_idx, column=1).border = BORDER

    ws.cell(row=row_idx, column=2, value=int(s['shift'])).font = DATA_FONT
    ws.cell(row=row_idx, column=2).alignment = CENTER
    ws.cell(row=row_idx, column=2).border = BORDER

    try:
        wh = int(s['wkly_hrs'])
    except (ValueError, TypeError):
        wh = s['wkly_hrs']
    ws.cell(row=row_idx, column=3, value=wh).font = DATA_FONT
    ws.cell(row=row_idx, column=3).alignment = CENTER
    ws.cell(row=row_idx, column=3).border = BORDER

    for col_idx, d in enumerate(['Sun', 'Mon', 'Tue', 'Wed', 'Thur', 'Fri', 'Sat'], 4):
        info = s['days'].get(d, {})
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = DATA_FONT
        cell.alignment = CENTER
        cell.border = BORDER

        if info.get('time') == 'OFF':
            cell.value = 'OFF'
            cell.fill = OFF_FILL
        elif not info.get('time'):
            cell.value = '--'
            cell.fill = NONE_FILL
        else:
            try:
                cell.value = int(info.get('hrs', 0))
            except (ValueError, TypeError):
                cell.value = info.get('hrs', '')

# Autofilter
ws.auto_filter.ref = f"A1:J{len(data)+1}"

wb.save('Employee Schedule 2026.xlsx')
print(f"Saved Employee Schedule 2026.xlsx with {len(data)} employees")
